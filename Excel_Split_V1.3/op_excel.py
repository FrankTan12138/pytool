#!/usr/bin/env python
# -*- coding: UTF-8 -*-


'''
@title: 对excel的操作
@author: Frank
@date: 2019/10/22

'''

#引入模块
import xlrd
import pandas as pd
import xlwt
import openpyxl
import urllib
from win32com.client import DispatchEx
import xlwings as xw
from PIL import ImageGrab
from xlutils.copy import copy as xl_copy
from urllib.request import unquote
from op_read_config import read_config  #读取配置文件
import traceback
import sys
import time

#读取Excel的sheetname名称
def read_sheet_name(os_path,file_name):
    excel_info=pd.read_excel(os_path+file_name,None)
    sheet_name_list=excel_info.keys()
    return(list(sheet_name_list))


#读取Excel数据
def read_excel(os_path,file_name,sheet_para,var_para,order_list):
    sheetnames=xlrd.open_workbook(os_path+file_name).sheet_names()   #打开来源数据excel表格
    data=pd.read_excel(os_path+file_name,sheet_name=sheetnames[int(sheet_para)],header= eval(var_para.replace("1","None")))   #有标题栏，读取sheet内的数据 
    if len(order_list) > 0:
        data=data[order_list]
    else:
        pass
    nrows=data.columns.size    #EXCEL列数
    ncols=len(data)
#     print(str(ncols) +"\t" +str(nrows))
    data_result={}
    for i in range(0,ncols):
        data_col=""
        for j in range(0,nrows):
            data_col=(data_col+","+"\'"+str(data.iloc[i,j]).replace("nan","").replace("\n","").replace(" ","")+"\'").strip(",")        
        data_result[i]="("+data_col+")"
    return(data_result)

#数据筛选
def data_filter(os_path,config_file,data_info,start_num,end_num,filter_condition,filter_name_num):
    para_name=read_config(os_path,config_file)
    filter_name_list=para_name['condition']['filter_name'].split(",")
    split_sheet_no=para_name['base']['split_sheet_no'].split(",")
    split_sheet=dict(zip(split_sheet_no,filter_name_list))
    data_result1={} #初始化
    j=1  #筛选结果重新序号
    for i in range(int(start_num)-1,end_num):
        data=data_info[i].replace("\'","").replace("(","").replace(")","").split(",")
        if i == 0:
            data_keys=data
            data_result1[i]=list(data)
        else:
            data_values=data
            data_result=dict(zip(data_keys,data_values))
            if str(filter_name_num) in split_sheet:
                filter_name=split_sheet[str(filter_name_num)]
                if data_result[filter_name] == filter_condition:
                    data_result1[j]=data_result
            else:
                data_result1[j]=data_result
                j=j+1
    return(data_result1)



#数据写入Excel
def write_excel(os_path,file_name,sheet_name,start_mun,insert_info):
    work_book=xlwt.Workbook(encoding='utf-8')
    sheet=work_book.add_sheet(sheet_name) #sheet名称
    #写入title数据
    for i in range(int(start_mun)-1,len(insert_info[0])):
        sheet.write(0,i,insert_info[0][i])
    #写入数据
        for j in range(int(start_mun),len(insert_info)):
            sheet.write(j,i,insert_info[j][insert_info[0][i]])
    #数据保存到excel   
    work_book.save(os_path+file_name+".xls")
    print("数据写入Excel-{}：sheet-{},执行完成!~".format(file_name,sheet_name))
    
    
#增加sheet写入Excel
def add_sheet(os_path,file_name,sheet_name,start_mun,insert_info):
    wb=xlrd.open_workbook(os_path+file_name+".xlsx")
    work_book = xl_copy(wb)
    sheet=work_book.add_sheet(sheet_name) #sheet名称
    #写入title数据
    for i in range(int(start_mun)-1,len(insert_info[0])):
        sheet.write(0,i,insert_info[0][i])
    #写入数据
        for j in range(int(start_mun),len(insert_info)):
            sheet.write(j,i,insert_info[j][insert_info[0][i]])
    #数据保存到excel   
    work_book.save(os_path+file_name+".xls")
    print("数据写入Excel-{}：sheet-{},执行完成!~".format(file_name,sheet_name))

#增加sheet写入Excel-xlsx版本   
def add_sheet_xlsx(os_path,file_name,sheet_name,start_mun,insert_info):
    wb = openpyxl.load_workbook(os_path+file_name+".xlsx")  #导入工作簿
    wb.create_sheet(sheet_name)  #增加一个sheet
    sheet = wb[sheet_name]  #打开刚创建的sheet_name
#写入title数据
    for i in range(int(start_mun)-1,len(insert_info[0])):
        sheet.cell(row=1,column=i+1).value=insert_info[0][i]
    #写入数据
        for j in range(int(start_mun),len(insert_info)):
            sheet.cell(row=j+1,column=i+1).value=insert_info[j][insert_info[0][i]]
    #数据保存到excel   
    wb.save(os_path+file_name+".xlsx")
    print("数据写入Excel-{}：sheet-{},执行完成!~".format(file_name,sheet_name))
    
#增加sheet写入Excel-xlsx版本(转置)
def add_sheet_xlsx_transfrom(os_path,file_name,sheet_name,start_mun,insert_info):
    wb = openpyxl.load_workbook(os_path+file_name+".xlsx")  #导入工作簿
    wb.create_sheet(sheet_name)  #增加一个sheet
    sheet = wb[sheet_name]  #打开刚创建的sheet_name
#写入title数据
    for i in range(int(start_mun)-1,len(insert_info[0])):
        sheet.cell(row=i+1,column=1).value=insert_info[0][i]
    #写入数据
        for j in range(int(start_mun),len(insert_info)):
            sheet.cell(row=i+1,column=j+1).value=insert_info[j][insert_info[0][i]]
    #数据保存到excel   
    wb.save(os_path+file_name+".xlsx")
    print("数据写入Excel-{}：sheet-{},执行完成!~".format(file_name,sheet_name))   


#对Excel数据表格截屏-win32
def excel_catch_screen(os_path,file_name, sheet_name, screen_area,picture_name):
    excel = DispatchEx("Excel.Application")  #启动excel
    excel.Visible = True  #可视化
    excel.DisplayAlerts = False  #是否显示警告
        
    wb = excel.Workbooks.Open(os_path+file_name)  #打开excel
    ws = wb.Sheets(sheet_name)  #选择sheet
    
    #拆分截频区域和图片名称
    screen_area=screen_area.split(",")
    picture_name=picture_name.split(",")
    

    #循环处理每个截图区域
    for i in range (0,len(screen_area)):
        ws.Range(screen_area[i]).CopyPicture()  #复制图片区域
        time.sleep(2)
        ws.Paste()  #粘贴       
        excel.Selection.ShapeRange.Name = picture_name[i]  #将刚刚选择的Shape重命名，避免与已有图片混淆
        ws.Shapes(picture_name[i]).Copy()  # 选择图片
        time.sleep(2)
        img = ImageGrab.grabclipboard()  # 获取剪贴板的图片数据
        img_name = picture_name[i]+ ".PNG" #生成图片的文件名
        img.save(os_path+img_name)  #保存图片
        time.sleep(1)
        
    wb.Close(SaveChanges=0)  #关闭工作薄，不保存
    excel.Quit()  #退出excel
    
#对Excel数据表格截屏-xlwings
def excel_catch_screen_xlwings(os_path,file_name, sheet_name, screen_area,picture_name):
    app=xw.App(visible=True,add_book=False) #使用xlwings的app启动
    wb = app.books.open(os_path+file_name)        #打开文件
    sheet=wb.sheets[sheet_name]                  #选定sheet
    
    #拆分截频区域和图片名称
    screen_area=screen_area.split(",")
    picture_name=picture_name.split(",")
    
    #循环处理每个截图区域
    for i in range (0,len(screen_area)):
        cell_data=sheet[screen_area[i]]
        cell_data.api.CopyPicture()                   # 复制图片区域      
        sheet.api.Paste()                       # 粘贴
        time.sleep(2)
        pic=sheet.pictures[i]                #当前图片
        pic.api.Copy()                          #复制图片
        img = ImageGrab.grabclipboard()         # 获取剪贴板的图片数据
        img.save(os_path+picture_name[i] + ".png")             #保存图片   
    wb.close()  #不保存，直接关闭
    app.quit()  #退出app

#下载数据文件
def download_excel(url,os_path,file_name):
    print(unquote(url))
    print("开始下载：{}".format(file_name))
    try:
        urllib.request.urlretrieve(url,os_path+file_name)  #下载文件，并保存到指定文件夹
    except:
        traceback.print_exc()
        sys.exit()