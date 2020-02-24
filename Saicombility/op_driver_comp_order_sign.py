#!/usr/bin/env python
# -*- coding: UTF-8 -*-


'''
@title: 司机订单标签月报
@author: Frank
@date: 2020/02/05

'''
#引入模块
import datetime,time,traceback,sys
sys.path.append(r'E:\software\eclipse\workspace\resource')
from op_read_config import read_config
from op_mysql import export_data
from op_excel import add_sheet_xlsx
import openpyxl,shutil



if __name__ == '__main__':
    os_path=r"D:\Code\para_config\\"
    config_file="driver_comp_order_sign_config.ini"
    today=str(datetime.date.today()) #取今天的日期
    
#每月1号执行    
    if today[8:10] == '01' :
        try:
            para_name=read_config(os_path,config_file) #读取配置信息
        #base参数
            v_date=str(datetime.date.today() - datetime.timedelta(days=int(para_name['base']['v_para_cnt'])))  #获取上个周日的日期
        #mysql参数
            db_ip=para_name['mysql']['db_ip']  #ip地址
            db_port=para_name['mysql']['db_port']  #端口号
            db_username=para_name['mysql']['db_username']  #账户
            db_password=para_name['mysql']['db_password']  #密码
            person_table_name=para_name['mysql']['person_table_name']  #清单表名
            city_name_list=para_name['mysql']['city_name'].split(",") #城市名称
            all_table_name=para_name['mysql']['all_table_name'] #汇总表名
            person_order_list=para_name['mysql']['person_order_list'].split(",")+[v_date[0:4]+str(i).zfill(2) for i in range(1,int(v_date[5:7])+1)]  #清单字段名称
            all_order_list=para_name['mysql']['all_order_list'].split(",")
        #excel参数
            excel_path=para_name['excel']['excel_path'] #模板路径
            list_sheet_name=para_name['excel']['list_sheet_name'].format(v_date[0:4]) #清单sheet名称
            total_sheet_name=para_name['excel']['total_sheet_name'] #汇总sheet名称
        
                        
            print("=====================\n开始进行操作，操作过程会持续一段时间，请稍后.....")
            for city_name in  city_name_list:
                person_var_sql=para_name['mysql']['var_sql'].format(person_table_name,city_name)
                all_var_sql=para_name['mysql']['var_sql'].format(all_table_name,city_name).replace(";","")+" and t.month_id='{}';".format(v_date[0:4]+v_date[5:7])
                excel_name=para_name['excel']['excel_name'].format(city_name) #模板名称
            #将mysql计算结果导入到Excel表格中-清单list
                mysql_list_data=export_data(db_ip,db_port,db_username,db_password,person_var_sql)
                list_data_info=mysql_list_data.data_filter("","",person_order_list)  #从mysql里导出清单数据 
                add_sheet_xlsx(excel_path,excel_name.split(".")[0],"list_temp",1,list_data_info)  #写入现有的excel增加sheet中
            #将mysql计算结果导入到Excel表格中-汇总list
                mysql_total_data=export_data(db_ip,db_port,db_username,db_password,all_var_sql)
                total_data_info=mysql_total_data.data_filter("","",all_order_list)  #从mysql里导出清单数据 
                add_sheet_xlsx(excel_path,excel_name.split(".")[0],"total_temp",1,total_data_info)  #写入现有的excel增加sheet中
            #数据写入excel-清单list
                wb = openpyxl.load_workbook(excel_path+excel_name)
                list_temp=wb['list_temp'] #清单临时sheet
                list_sheet=wb[list_sheet_name]  #清单数据
                row_num=list_temp.max_column #列数
                col_num=list_temp.max_row #行数
                k=1 #初始值
            #覆盖掉清单sheet对应的数据
                for i in range(1,list_temp.max_column+1):
                    if i != 3:  #剔除年份
                        for j in range(2,list_temp.max_row+1): #从第一行数据读取
                            if list_temp.cell(row=j,column=i).value ==' None' :
                                list_sheet.cell(row=j,column=k).value=''
                            else:
                                list_sheet.cell(row=j,column=k).value=list_temp.cell(row=j,column=i).value #覆盖掉清单list的数据
                        k=k+1
                wb.remove(list_temp) #删除临时新增的list_temp
                print("数据写入Excel-{}：sheet-{},执行完成!~".format(excel_name,list_sheet_name))
            #数据写入excel-汇总list
                total_temp=wb['total_temp'] #汇总临时sheet
                total_sheet=wb[total_sheet_name]  #汇总数据
                total_col_num=total_sheet.max_row #汇总行数
                temp_col_num=total_temp.max_row  #临时行数
                temp_row_num=total_temp.max_column  #临时行数
                total_sheet.cell(row=total_col_num+1,column=2).value=v_date[0:4]+v_date[5:7] #日期
                total_sum=0 #初始值
                for i in range(2,temp_col_num+1) :
                    if total_temp.cell(row=i,column=3).value == ' 高频' :
                        total_sheet.cell(row=total_col_num+1,column=3).value=total_temp.cell(row=i,column=4).value
                        total_sum=total_sum+int(total_temp.cell(row=i,column=4).value)
                    elif total_temp.cell(row=i,column=3).value == ' 沉默' :
                        total_sheet.cell(row=total_col_num+1,column=6).value=total_temp.cell(row=i,column=4).value
                        total_sum=total_sum+int(total_temp.cell(row=i,column=4).value)
                    elif total_temp.cell(row=i,column=3).value == ' 无完单':
                        total_sheet.cell(row=total_col_num+1,column=7).value=total_temp.cell(row=i,column=4).value
                        total_sum=total_sum+int(total_temp.cell(row=i,column=4).value)
                    elif total_temp.cell(row=i,column=3).value == ' 低频' :
                        total_sheet.cell(row=total_col_num+1,column=5).value=total_temp.cell(row=i,column=4).value
                        total_sum=total_sum+int(total_temp.cell(row=i,column=4).value)
                    elif total_temp.cell(row=i,column=3).value == ' 中频' :
                        total_sheet.cell(row=total_col_num+1,column=4).value=total_temp.cell(row=i,column=4).value
                        total_sum=total_sum+int(total_temp.cell(row=i,column=4).value)
                    else:
                        pass
                total_sheet.cell(row=total_col_num+1,column=8).value=total_sum #汇总
                wb.remove(total_temp) #删除临时新增的list_temp
                wb.save(excel_path+excel_name)
                print("数据写入Excel-{}：sheet-{},执行完成!~".format(excel_name,total_sheet_name))
            #将模板的数据存入新的excel里    
                shutil.copyfile(excel_path+excel_name, excel_path+excel_name.split(".")[0]+"_"+v_date[0:4]+v_date[5:7]+".xlsx")
                print("拷贝Excel-{},执行完成!~".format(excel_name.split(".")[0]+"_"+v_date[0:4]+v_date[5:7]+".xlsx"))
                
            print("=====================\n操作结束，界面会在数秒后自动关闭...")
            time.sleep(3)
            
        except:
            traceback.print_exc()
            sys.exit()
