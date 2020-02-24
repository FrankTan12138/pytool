#!/usr/bin/env python
# -*- coding: UTF-8 -*-


'''
@title: 全国运力日报
@author: Frank
@date: 2020/01/16

'''
#引入模块
from op_read_config import read_config
from op_mysql import export_data
from op_excel import excel_catch_screen, add_sheet_xlsx_transfrom
from op_zipfile import op_sftp
from op_send_mail import send_mail
from op_dingtalk import dingtalk_chatbot
import datetime,time,traceback,sys,os
import openpyxl,shutil


if __name__ == '__main__':
    os_path=r"D:\Code\para_config\\"
    config_file="dingtalk_yunli_daily_config.ini"
    
    
    try:
        para_name=read_config(os_path,config_file) #读取配置信息
    #base参数
        today=str(datetime.date.today()) #根据updatetime从mysql中取要求的数据
        v_date=str(datetime.date.today() - datetime.timedelta(days=int(para_name['base']['v_para_cnt'])))  #获取日期
    #Mysql参数
        db_ip=para_name['mysql']['db_ip']  #ip地址
        db_port=int(para_name['mysql']['db_port'])  #端口
        db_username=para_name['mysql']['db_username']  #账户
        db_password=para_name['mysql']['db_password']  #密码
        order_list=para_name['mysql']['order_list'].split(",")  #字段名称
        var_sql=para_name['mysql']['var_sql'].format(para_name['mysql']['table_name'],today)
        
    #excel部分参数
        excel_path=para_name['excel']['excel_path'] #excel路径
        excel_name=para_name['excel']['excel_name']  #excel名称
        sheet_name=para_name['excel']['sheet_name']  #sheet名称
        screen_area=para_name['excel']['screen_area'] #截屏区域，多个用逗号隔开
        picture_name=para_name['excel']['picture_name'].format(v_date.replace("-",""))
    #图片生成链接部分
        host_ip=para_name['aliyun']['host_ip'] #ip地址
        port=int(para_name['aliyun']['port']) #端口
        username=para_name['aliyun']['username']  #登录账户
        password=para_name['aliyun']['password']  #登录密码
        remote_path=para_name['aliyun']['remote_path']  #远程路径
    #dingtalk参数
        webhook=para_name['dingtalk']['webhook'].format(para_name['dingtalk']['access_token']).replace("\"","")  #api接口
        headers=para_name['dingtalk']['headers']  #headers信息
    #markdwon参数
        title=para_name['markdown']['title']  #title信息
        text=para_name['markdown']['text'].format(v_date,v_date.replace("-",""),time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))).replace("\\n","\n") #msg信息内容
        is_at_all=para_name['markdown']['is_at_all']  #是否@所有人
    #邮箱参数
        host_server=para_name['mail']['host_server']  #邮箱服务器ip
        host_port=para_name['mail']['host_port']  #端口
        mail_username=para_name['mail']['user_name']  #登录账户
        mail_password=para_name['mail']['password']  #密码
        sender=para_name['mail']['sender']  #发件邮箱
        receiver=para_name['mail']['receiver'] #收件邮箱
        cc=para_name['mail']['cc']  #抄送
        mail_title=para_name['mail']['mail_title'].format(v_date[5:7]+v_date[8:10])  #标题
        mail_content=para_name['mail']['mail_content'].format(v_date[0:4],v_date[5:7],v_date[8:10]) #内容
        attachment_img=para_name['mail']['attachment_img'] #图片
        attachment_txt=para_name['mail']['attachment_txt'] #txt
        attachment_pdf=para_name['mail']['attachment_pdf'] #pdf
        attachment_excel=excel_path+excel_name.split(".")[0]+"_"+v_date.replace("-","")+".xlsx" #excel
        attachment_word=para_name['mail']['attachment_word'] #word
    
        print("=====================\n开始进行操作，操作过程会持续一段时间，请稍后.....")
    #将mysql计算结果导入到Excel表格中
        data_result={} #初始化
        mysql_data=export_data(db_ip,db_port,db_username,db_password,var_sql)
        data_info=mysql_data.data_filter("","",order_list)  #从mysql里导出数据
        add_sheet_xlsx_transfrom(excel_path,excel_name.split(".")[0],"temp",1,data_info)  #转置并写入现有的excel增加sheet中
    #覆盖掉input对应的数据
        wb = openpyxl.load_workbook(excel_path+excel_name)
        sheet1=wb['input']  #选中sheet-input
        sheet2=wb['temp'] #选中刚导入Excel的temp
        for i in range(1,sheet1.max_column+1):
            for j in range(1,sheet1.max_row+1):
                sheet1.cell(row=j,column=i).value=sheet2.cell(row=j,column=i).value #覆盖掉sheet-input的数据
        wb.remove(sheet2) #删除临时新增的sheet1
        wb.save(excel_path+excel_name)
        print("数据加载到Excel-享道出行-全国运力日报模板：sheet-{},执行完成!~".format('input'))
    #将模板的数据存入新的excel里    
        shutil.copyfile(excel_path+excel_name, excel_path+excel_name.split(".")[0]+"_"+v_date.replace("-","")+".xlsx")        
    #将模板的数据覆盖掉新excel里的数据  
        wb1 = openpyxl.load_workbook(excel_path+excel_name.split(".")[0]+"_"+v_date.replace("-","")+".xlsx")
        wb1['input'].sheet_state = 'hidden'  #隐藏掉sheet-input
        wb1.save(excel_path+excel_name.split(".")[0]+"_"+v_date.replace("-","")+".xlsx")
        print("目标Excel：享道出行-全国运力日报_{},执行完成!~".format(v_date.replace("-","")))
         
    #将Excel数据截屏
        excel_catch_screen(excel_path,excel_name.split(".")[0]+"_"+v_date.replace("-","")+".xlsx", sheet_name, screen_area,picture_name)
        print("Excel：享道出行-全国运力日报_{},数据截屏执行完成！~".format(v_date.replace("-","")))
         
    #图片上传到文件夹
        sftp=op_sftp(host_ip,port,username,password)
        sftp.sftp_put(excel_path,picture_name+".PNG",remote_path,picture_name+".PNG")
        print("图片：享道出行-全国运力日报_{},上传到文件夹执行完成！~".format(v_date.replace("-","")))        
          
    #通过钉钉发送Markdown
        dingtalk_chatbot=dingtalk_chatbot(webhook)
        dingtalk_chatbot.Dingtalk_markdown(title,text,is_at_all)  #推送信息
        print("钉钉推送信息:享道出行-全国运力日报_{},执行完成！~".format(v_date.replace("-","")))

    #发送邮件
        send_mail(host_server,host_port,mail_username,mail_password,sender,receiver,cc,mail_title,mail_content,attachment_img,attachment_txt,attachment_pdf,attachment_excel,attachment_word)
        print("邮件:享道出行-全国运力日报_{},执行完成！~".format(v_date.replace("-",""))) 

        
        os.remove(excel_path+picture_name+".PNG")  #删除图片
        os.remove(excel_path+excel_name.split(".")[0]+"_"+v_date.replace("-","")+".xlsx") #删除邮件附件
        print("=====================\n操作结束，界面会在数秒后自动关闭...")
        time.sleep(3)
        
    except:
        traceback.print_exc()
        sys.exit()
        