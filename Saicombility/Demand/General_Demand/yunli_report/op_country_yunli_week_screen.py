#!/usr/bin/env python
# -*- coding: UTF-8 -*-


'''
@title: 司机收入周报处理
@author: Frank
@date: 2020/01/27

'''

#引入模块
import datetime,time,traceback,sys
sys.path.append(r'E:\software\eclipse\workspace\resource')
from op_read_config import read_config
from op_mysql import export_data
from op_excel import excel_catch_screen, add_sheet_xlsx_transfrom
from op_zipfile import op_sftp
import openpyxl,shutil

if __name__ == '__main__':
    os_path=r"D:\Code\para_config\\"
    config_file="dingtalk_yunli_week_config.ini"
    week_id=datetime.date.today().strftime("%w")  #今天是星期几
    
    if week_id == '1' :   #周次才发送周报    
        try:
            para_name=read_config(os_path,config_file) #读取配置信息
        #base参数       
            v_date=str(datetime.date.today() - datetime.timedelta(days=int(para_name['base']['v_para_cnt'])))  #获取上个周日的日期
            v_week_num=str(datetime.datetime.strptime(v_date,'%Y-%m-%d').isocalendar()[0])+str(datetime.datetime.strptime(v_date,'%Y-%m-%d').isocalendar()[1]).zfill(2)  #本年第几周
        #Mysql参数
            db_ip=para_name['mysql']['db_ip']  #ip地址
            db_port=int(para_name['mysql']['db_port'])  #端口
            db_username=para_name['mysql']['db_username']  #账户
            db_password=para_name['mysql']['db_password']  #密码
            order_list=para_name['mysql']['order_list'].split(",")  #字段名称
            var_sql=para_name['mysql']['var_sql'].format(para_name['mysql']['table_name'])       
        #excel部分参数
            excel_path=para_name['excel']['excel_path'] #excel路径
            excel_name=para_name['excel']['excel_name']  #excel名称
            sheet_name=para_name['excel']['sheet_name']  #sheet名称
            screen_area=para_name['excel']['screen_area'] #截屏区域，多个用逗号隔开
            picture_name=para_name['excel']['picture_name'].format(v_week_num)
        #图片生成链接部分
            host_ip=para_name['aliyun']['host_ip'] #ip地址
            port=int(para_name['aliyun']['port']) #端口
            username=para_name['aliyun']['username']  #登录账户
            password=para_name['aliyun']['password']  #登录密码
            remote_path=para_name['aliyun']['remote_path']  #远程路径
          
            
            print("=====================\n开始进行操作，操作过程会持续一段时间，请稍后.....")
        #将mysql计算结果导入到Excel表格中
            data_result={} #初始化
            mysql_data=export_data(db_ip,db_port,db_username,db_password,var_sql)
            data_info=mysql_data.data_filter("","",order_list)  #从mysql里导出数据
            add_sheet_xlsx_transfrom(excel_path,excel_name.split(".")[0],"temp",1,data_info)  #转置并写入现有的excel增加sheet中
        #覆盖掉input对应的数据
            wb = openpyxl.load_workbook(excel_path+excel_name)
            sheet1=wb['input']  #选中sheet-input
            sheet2=wb['temp'] #选中刚导入Excel的temp
            for i in range(1,sheet1.max_column+1):
                for j in range(1,sheet1.max_row+1):
                    sheet1.cell(row=j,column=i).value=sheet2.cell(row=j,column=i).value #覆盖掉sheet-input的数据
            wb.remove(sheet2) #删除临时新增的sheet1
            wb.save(excel_path+excel_name)
            print("数据加载到Excel-享道出行-全国运力周报模板：sheet-{},执行完成!~".format('input'))    
        #将模板的数据存入新的excel里    
            shutil.copyfile(excel_path+excel_name, excel_path+excel_name.split(".")[0]+"_"+v_week_num+".xlsx")        
        #将模板的数据覆盖掉新excel里的数据  
            wb1 = openpyxl.load_workbook(excel_path+excel_name.split(".")[0]+"_"+v_week_num+".xlsx")
            wb1['input'].sheet_state = 'hidden'  #隐藏掉sheet-input
            wb1.save(excel_path+excel_name.split(".")[0]+"_"+v_week_num+".xlsx")
            print("目标Excel：享道出行-全国运力周报_{},执行完成!~".format(v_week_num))
             
        #将Excel数据截屏
            excel_catch_screen(excel_path,excel_name.split(".")[0]+"_"+v_week_num+".xlsx", sheet_name, screen_area,picture_name)
            print("Excel：享道出行-全国运力周报_{},数据截屏执行完成！~".format(v_week_num))
            
        #图片上传到文件夹
            sftp=op_sftp(host_ip,port,username,password)
            sftp.sftp_put(excel_path,picture_name+".PNG",remote_path,picture_name+".PNG")
            print("图片：享道出行-全国运力周报_{},上传到文件夹执行完成！~".format(v_week_num))  
            
            print("=====================\n操作结束，界面会在数秒后自动关闭...")
            time.sleep(3)
        except:
            traceback.print_exc()
            sys.exit()
    else:
        print("今天是周{}，无需发送周报!~".format(week_id))
        sys.exit()